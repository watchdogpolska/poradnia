import json

from django.conf import settings
from django.contrib import admin
from django.http import HttpResponse
from django.utils import timezone
from django.utils.html import format_html
from django.utils.translation import gettext_lazy as _
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from poradnia.advicer.models import Area, InstitutionKind, Issue, PersonKind
from poradnia.teryt.models import JST

from .models import N8nArticlesSearchRequest, N8nCaseTagsRequest

HEADER_FONT = Font(bold=True)
HEADER_FILL = PatternFill(
    start_color="FFDDDDDD", end_color="FFDDDDDD", fill_type="solid"
)
STANDARD_COLUMN_WIDTH = 8.43


def _naive(dt):
    return timezone.localtime(dt).replace(tzinfo=None) if dt else ""


def _parse_ai_response(response_text):
    if not response_text:
        return {}
    try:
        return json.loads(response_text)
    except (TypeError, ValueError):
        return {}


def _phrase_matches_pretty(phrase_matches):
    if not phrase_matches:
        return ""
    data = [
        {k: v for k, v in item.items() if k != "summary"} for item in phrase_matches
    ]
    return json.dumps(data, indent=2, ensure_ascii=False)


@admin.register(N8nArticlesSearchRequest)
class N8nArticlesSearchRequestAdmin(admin.ModelAdmin):
    list_display = (
        "id",
        "request_id",
        "environment",
        "status",
        "is_foi",
        "direct_search",
        "case",
        "letter",
        "accepted_by",
        "rejected_by",
        "created_at",
        "updated_at",
    )
    list_filter = ("environment", "status", "is_foi", "direct_search")
    search_fields = ("request_id", "question", "response")
    date_hierarchy = "created_at"
    raw_id_fields = ("case", "letter")
    exclude = ("phrase_matches",)
    actions = ["export_as_excel"]

    def get_readonly_fields(self, request, obj=None):
        fields = [f.name for f in self.model._meta.fields]
        fields[fields.index("phrase_matches")] = "phrase_matches_display"
        return fields

    def has_add_permission(self, request):
        return False

    def has_change_permission(self, request, obj=None):
        return False

    @admin.display(description=_("Phrase matches"))
    def phrase_matches_display(self, obj):
        if not obj.phrase_matches:
            return "-"
        pretty = json.dumps(obj.phrase_matches, indent=2, ensure_ascii=False)
        return format_html("<pre>{}</pre>", pretty)

    @admin.action(description=_("Export selected as Excel report"))
    def export_as_excel(self, request, queryset):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "N8nArticlesSearchRequest"

        columns = [
            "id",
            "created_at",
            "request_id",
            "environment",
            "status",
            "is_foi",
            "direct_search",
            "case_id",
            "case_absolute_url",
            "letter_id",
            "letter_absolute_url",
            "accepted_by",
            "rejected_by",
            "rejection_reason",
            "response",
            "phrase_matches",
        ]
        sheet.append(columns)
        for col_idx in range(1, len(columns) + 1):
            cell = sheet.cell(row=1, column=col_idx)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL

        case_url_col = columns.index("case_absolute_url") + 1
        letter_url_col = columns.index("letter_absolute_url") + 1
        wrap_col_start = columns.index("rejection_reason") + 1
        for col_idx in range(wrap_col_start, len(columns) + 1):
            sheet.column_dimensions[get_column_letter(col_idx)].width = (
                STANDARD_COLUMN_WIDTH * 3
            )

        queryset = queryset.select_related(
            "case", "letter", "accepted_by", "rejected_by"
        )
        for obj in queryset:
            case_url = (
                request.build_absolute_uri(obj.case.get_absolute_url())
                if obj.case_id
                else ""
            )
            letter_url = (
                request.build_absolute_uri(obj.letter.get_absolute_url())
                if obj.letter_id
                else ""
            )
            sheet.append(
                [
                    obj.id,
                    _naive(obj.created_at),
                    obj.request_id,
                    obj.environment,
                    obj.status,
                    obj.is_foi,
                    obj.direct_search,
                    obj.case_id,
                    case_url,
                    obj.letter_id,
                    letter_url,
                    str(obj.accepted_by) if obj.accepted_by_id else "",
                    str(obj.rejected_by) if obj.rejected_by_id else "",
                    obj.rejection_reason or "",
                    obj.response,
                    _phrase_matches_pretty(obj.phrase_matches),
                ]
            )
            row_idx = sheet.max_row
            for col_idx in range(wrap_col_start, len(columns) + 1):
                sheet.cell(row=row_idx, column=col_idx).alignment = Alignment(
                    wrap_text=True, vertical="top"
                )
            for url, col_idx in (
                (case_url, case_url_col),
                (letter_url, letter_url_col),
            ):
                if url:
                    cell = sheet.cell(row=row_idx, column=col_idx)
                    cell.hyperlink = url
                    cell.style = "Hyperlink"

        sheet.auto_filter.ref = sheet.dimensions

        response = HttpResponse(
            content_type=(
                "application/vnd.openxmlformats-officedocument" ".spreadsheetml.sheet"
            )
        )
        timestamp = timezone.localtime(timezone.now()).strftime("%Y%m%d_%H%M%S")
        filename = f"n8n_articles_search_requests_{settings.APP_MODE}_{timestamp}.xlsx"
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        workbook.save(response)
        return response


@admin.register(N8nCaseTagsRequest)
class N8nCaseTagsRequestAdmin(admin.ModelAdmin):
    list_display = (
        "id",
        "request_id",
        "environment",
        "status",
        "case",
        "accepted_by",
        "rejected_by",
        "created_at",
        "updated_at",
    )
    list_filter = ("environment", "status")
    search_fields = ("request_id", "question", "response")
    date_hierarchy = "created_at"
    raw_id_fields = ("case",)
    actions = ["export_as_excel"]

    def get_readonly_fields(self, request, obj=None):
        return [f.name for f in self.model._meta.fields]

    def has_add_permission(self, request):
        return False

    def has_change_permission(self, request, obj=None):
        return False

    @admin.action(description=_("Export selected as Excel report"))
    def export_as_excel(self, request, queryset):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "N8nCaseTagsRequest"

        columns = [
            "id",
            "created_at",
            "request_id",
            "environment",
            "status",
            "case_id",
            "case_absolute_url",
            "advice_id",
            "advice_absolute_url",
            "accepted_by",
            "rejected_by",
            "rejection_reason",
            "advice_issues_names",
            "ai_response_issues_names",
            "advice_areas_names",
            "ai_response_areas_names",
            "advice_institutionkind_name",
            "ai_response_institutionkind_name",
            "advice_personkind_name",
            "ai_response_personkind_name",
            "advice_jst_name",
            "ai_response_jst_name",
            "advice_subject",
            "ai_response_subject",
            "ai_response_summary",
        ]
        sheet.append(columns)
        for col_idx in range(1, len(columns) + 1):
            cell = sheet.cell(row=1, column=col_idx)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL

        case_url_col = columns.index("case_absolute_url") + 1
        advice_url_col = columns.index("advice_absolute_url") + 1
        wrap_col_start = columns.index("rejection_reason") + 1
        for col_idx in range(wrap_col_start, len(columns) + 1):
            sheet.column_dimensions[get_column_letter(col_idx)].width = (
                STANDARD_COLUMN_WIDTH * 3
            )

        issue_names = dict(Issue.objects.values_list("id", "name"))
        area_names = dict(Area.objects.values_list("id", "name"))
        person_kind_names = dict(PersonKind.objects.values_list("id", "name"))
        institution_kind_names = dict(InstitutionKind.objects.values_list("id", "name"))

        queryset = queryset.select_related(
            "case__advice__institution_kind",
            "case__advice__person_kind",
            "case__advice__jst__parent__parent",
            "accepted_by",
            "rejected_by",
        ).prefetch_related("case__advice__issues", "case__advice__area")

        jst_name_cache = {}

        def _ai_jst_full_name(jst_id):
            if jst_id not in jst_name_cache:
                jst = (
                    JST.objects.select_related("parent__parent")
                    .filter(pk=jst_id)
                    .first()
                )
                jst_name_cache[jst_id] = str(jst) if jst else ""
            return jst_name_cache[jst_id]

        for obj in queryset:
            case_url = (
                request.build_absolute_uri(obj.case.get_absolute_url())
                if obj.case_id
                else ""
            )
            advice = getattr(obj.case, "advice", None) if obj.case_id else None
            advice_url = (
                request.build_absolute_uri(advice.get_absolute_url()) if advice else ""
            )

            ai_data = _parse_ai_response(obj.response)
            ai_issue_ids = ai_data.get("issues") or []
            ai_area_ids = ai_data.get("area") or []
            ai_institution_kind_id = ai_data.get("institution_kind")
            ai_person_kind_id = ai_data.get("person_kind")
            ai_jst_id = ai_data.get("jst")

            sheet.append(
                [
                    obj.id,
                    _naive(obj.created_at),
                    obj.request_id,
                    obj.environment,
                    obj.status,
                    obj.case_id,
                    case_url,
                    advice.pk if advice else "",
                    advice_url,
                    str(obj.accepted_by) if obj.accepted_by_id else "",
                    str(obj.rejected_by) if obj.rejected_by_id else "",
                    obj.rejection_reason or "",
                    (
                        ", ".join(advice.issues.values_list("name", flat=True))
                        if advice
                        else ""
                    ),
                    ", ".join(issue_names[i] for i in ai_issue_ids if i in issue_names),
                    (
                        ", ".join(advice.area.values_list("name", flat=True))
                        if advice
                        else ""
                    ),
                    ", ".join(area_names[i] for i in ai_area_ids if i in area_names),
                    (
                        advice.institution_kind.name
                        if advice and advice.institution_kind_id
                        else ""
                    ),
                    institution_kind_names.get(ai_institution_kind_id, ""),
                    (
                        advice.person_kind.name
                        if advice and advice.person_kind_id
                        else ""
                    ),
                    person_kind_names.get(ai_person_kind_id, ""),
                    str(advice.jst) if advice and advice.jst_id else "",
                    _ai_jst_full_name(ai_jst_id) if ai_jst_id else "",
                    advice.subject if advice and advice.subject else "",
                    ai_data.get("subject") or "",
                    ai_data.get("summary") or "",
                ]
            )
            row_idx = sheet.max_row
            for col_idx in range(wrap_col_start, len(columns) + 1):
                sheet.cell(row=row_idx, column=col_idx).alignment = Alignment(
                    wrap_text=True, vertical="top"
                )
            for url, col_idx in (
                (case_url, case_url_col),
                (advice_url, advice_url_col),
            ):
                if url:
                    cell = sheet.cell(row=row_idx, column=col_idx)
                    cell.hyperlink = url
                    cell.style = "Hyperlink"

        sheet.auto_filter.ref = sheet.dimensions

        response = HttpResponse(
            content_type=(
                "application/vnd.openxmlformats-officedocument" ".spreadsheetml.sheet"
            )
        )
        timestamp = timezone.localtime(timezone.now()).strftime("%Y%m%d_%H%M%S")
        filename = f"n8n_case_tags_requests_{settings.APP_MODE}_{timestamp}.xlsx"
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        workbook.save(response)
        return response
