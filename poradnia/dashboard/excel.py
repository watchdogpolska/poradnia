import re

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from . import reports

INVALID_SHEET_CHARS = re.compile(r"[\[\]:*?/\\]")
MAX_SHEET_TITLE_LEN = 31

TITLE_FONT = Font(bold=True, size=13)
DESCRIPTION_FONT = Font(italic=True, color="FF666666")
HEADER_FONT = Font(bold=True)
HEADER_FILL = PatternFill(
    start_color="FFDDDDDD", end_color="FFDDDDDD", fill_type="solid"
)
PINNED_FONT = Font(bold=True)
PINNED_FILL = PatternFill(
    start_color="FFF5F5F5", end_color="FFF5F5F5", fill_type="solid"
)


def _unique_sheet_title(title, used_titles):
    base = INVALID_SHEET_CHARS.sub(" ", str(title)).strip() or "Sheet"
    base = base[:MAX_SHEET_TITLE_LEN]
    name = base
    suffix = 2
    while name.lower() in used_titles:
        tail = f" ({suffix})"
        name = base[: MAX_SHEET_TITLE_LEN - len(tail)] + tail
        suffix += 1
    used_titles.add(name.lower())
    return name


def _write_report_sheet(workbook, report, used_titles):
    sheet = workbook.create_sheet(_unique_sheet_title(report["title"], used_titles))
    columns = report["columns"]
    n_cols = max(len(columns), 1)

    sheet.append([str(report["title"])])
    sheet.cell(row=1, column=1).font = TITLE_FONT
    if n_cols > 1:
        sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)

    description = report.get("description")
    if description:
        sheet.append([str(description)])
        cell = sheet.cell(row=2, column=1)
        cell.font = DESCRIPTION_FONT
        cell.alignment = Alignment(wrap_text=True)
        if n_cols > 1:
            sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n_cols)

    header_row = sheet.max_row + 1
    sheet.append([str(col["label"]) for col in columns])
    for col_idx in range(1, n_cols + 1):
        cell = sheet.cell(row=header_row, column=col_idx)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL

    for row in report["rows"]:
        sheet.append([row.get(col["key"], "") for col in columns])
        if row.get("pinned"):
            row_idx = sheet.max_row
            for col_idx in range(1, n_cols + 1):
                cell = sheet.cell(row=row_idx, column=col_idx)
                cell.font = PINNED_FONT
                cell.fill = PINNED_FILL

    for col_idx, col in enumerate(columns, start=1):
        width = len(str(col["label"]))
        for row in report["rows"]:
            width = max(width, len(str(row.get(col["key"], ""))))
        sheet.column_dimensions[get_column_letter(col_idx)].width = min(
            max(width + 2, 10), 60
        )

    return sheet


def build_workbook(year):
    """One worksheet per dashboard report table (title + description + data),
    covering all 4 tabs. Summary and Courts are year-agnostic; Cases and Tags
    both use the single `year` passed in - the caller is responsible for
    resolving any mismatch between the Cases/Tags tab selectors beforehand.
    """
    workbook = Workbook()
    workbook.remove(workbook.active)
    used_titles = set()

    report_list = [
        reports.summary_report(),
        reports.users_report(year),
        reports.cases_report(year),
        reports.staff_letters_report(year),
        reports.issues_report(year),
        reports.areas_report(year),
        reports.institution_kind_report(year),
        reports.person_kind_report(year),
        reports.courtsessions_report(),
        reports.courtcase_report(),
    ]
    for report in report_list:
        _write_report_sheet(workbook, report, used_titles)

    return workbook
